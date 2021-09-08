import os
import urllib
import uuid # to create random filename

import bs4
import openpyxl
import requests
import pandas as pd
import threading
import matplotlib.pyplot as plt
import re

header = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,"
              "application/signed-exchange;v=b3;q=0.9",
    # "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "zh-CN,zh;q=0.9,en-GB;q=0.8,en;q=0.7",
    "Cache-Control": "max-age=0",
    "Connection": "keep-alive",
    "Host": "movie.douban.com",
    "sec-ch-ua": "\"Chromium\";v=\"92\", \" Not A;Brand\";v=\"99\", \"Google Chrome\";v=\"92\"",
    "sec-ch-ua-mobile": "?0",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
    "Upgrade-Insecure-Requests": "1",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/92.0.4515.159 Safari/537.36 "
}


class Movie:
    def __init__(self):
        self.name = ""
        self.director = ""
        self.star = ""
        self.year = ""
        self.mtype = ""
        self.score = ""
        self.comment = ""
        self.img_link = ""

    def update(self, name="", director="", star="", year="", mtype="", score="", comment="", img_link=""):
        self.name = name
        self.director = director
        self.star = star
        self.year = year
        self.mtype = mtype
        self.score = score
        self.comment = comment
        self.img_link = img_link
# actually, the function Movie.update can be added to the struction function"__init__()" 


def get25movie(url):  # get 25 movies info by the given url and save them to the list "movies"
    global movies
    print(url)
    res = requests.get(url, headers=header)
    html = res.text
    soup = bs4.BeautifulSoup(html, features="lxml")
    elems = soup.find("ol", attrs={"class": "grid_view"}).find_all("li")
    for elem in elems:
        img_link = elem.find("div", attrs={"class": "pic"}).find("img")["src"]
        name = elem.find("div", attrs={"class": "info"}).find("span", attrs={"class": "title"}).text
        # print(elem.find("div", attrs={"class": "bd"}).find("p").text.split("\n")[1].strip().split("主演: "))
        director = elem.find("div", attrs={"class": "bd"}).find("p").text.split("\n")[1].strip().split("主演: ")[
            0].replace(
            "导演: ", "")
        try:
            star = elem.find("div", attrs={"class": "bd"}).find("p").text.split("\n")[1].strip().split("主演: ")[1]
        except Exception as error:
            star = ""
        year = \
            elem.find("div", attrs={"class": "bd"}).find("p").text.split("\n")[2].strip().replace("\xa0", "").split(
                "/")[0].replace("(中国大陆)", "")
        mtype = \
            elem.find("div", attrs={"class": "bd"}).find("p").text.split("\n")[2].strip().replace("\xa0", "").split(
                "/")[2]
        score = elem.find("div", attrs={"class": "bd"}).find("span", attrs={"class": "rating_num"}).text
        try:
            comment = elem.find("div", attrs={"class": "bd"}).find("p", attrs={"class": "quote"}).text.strip()
        except Exception as error:
            comment = ""
        movie = Movie()
        movie.update(name, director, star, year, mtype, score, comment, img_link)
        movies.append(movie)


def save_xlsx():  # save the movies to .xlsx file
    wb = openpyxl.Workbook()
    sheet = wb["Sheet"]
    global movies
    sheet.cell(row=1, column=1).value = "电影名"
    sheet.cell(row=1, column=2).value = "导演"
    sheet.cell(row=1, column=3).value = "主演"
    sheet.cell(row=1, column=4).value = "年份"
    sheet.cell(row=1, column=5).value = "电影类型"
    sheet.cell(row=1, column=6).value = "评分"
    sheet.cell(row=1, column=7).value = "评语"
    sheet.cell(row=1, column=8).value = "电影海报"
    for i in range(0, len(movies)):
        movie = movies[i]
        sheet.cell(row=i + 2, column=1).value = movie.name
        sheet.cell(row=i + 2, column=2).value = movie.director
        sheet.cell(row=i + 2, column=3).value = movie.star
        sheet.cell(row=i + 2, column=4).value = movie.year
        sheet.cell(row=i + 2, column=5).value = movie.mtype
        sheet.cell(row=i + 2, column=6).value = movie.score
        sheet.cell(row=i + 2, column=7).value = movie.comment
        sheet.cell(row=i + 2, column=8).value = movie.img_link
    wb.save("DouBan_movies.xlsx")


def trans_xlsx2csv():  # transform the sheet of excel to .csv file
    data = pd.read_excel('DouBan_movies.xlsx', 'Sheet', index_col=0)
    data.to_csv('data.csv', encoding='utf-8')


def download_img():  # use multithreading to download the img
    global movies
    my_movies = movies
    os.chdir(os.getcwd())  # 获取并进入当前文档所在路径
    try:
        os.mkdir("Pictures")
    except Exception as error:
        print(error)
    os.chdir("./Pictures")
    while True:
        glock.acquire()

        if len(my_movies) == 0:
            glock.release()
            break
        else:
            my_movie = my_movies.pop()
            url = my_movie.img_link
            name = my_movie.name
            glock.release()
            # 修改文件名
            path = '%s.png' % name
            print(path)
            # 下载图片，保存本地
            try:
                urllib.request.urlretrieve(url, filename=path)
            except Exception as error:
                print("电影" + name + "的电影海报下载失败")


def draw_plt():  # 绘制柱状图
    wb = openpyxl.load_workbook("DouBan_movies.xlsx")
    sheet = wb["Sheet"]
    temps = []
    for row in range(2, 252):
        temps.append([sheet.cell(row=row, column=4).value, sheet.cell(row=row, column=1).value])
    # print(temps)
    lists = dict()  # make a dict to save the year and the num of movies in that year
    for temp in temps:
        if temp[0] in lists:
            lists[temp[0]] += 1
        else:
            lists[temp[0]] = 1
    """for key, value in list(lists.items()):
        print(key, value)"""
    data = list()
    labels = list()
    for key, value in list(lists.items()):
        data.append(value)
        labels.append(key)
    plt.bar(range(len(data)), data, tick_label=labels)
    plt.show()
    # Attentions! 
    # the labels is unsorted because the dict cannot be sorted.
    #you can transform it to [[a,b]] and begin sort.

if __name__ == "__main__":
    movies = []
    glock = threading.Lock()
    # get the all URLs needed
    urls = ["https://movie.douban.com/top250?start=" + str(25 * i) for i in range(0, 10)]
    for link in urls:
        get25movie(link)
    print(len(movies))
    save_xlsx()
    trans_xlsx2csv()
    download_img()
    draw_plt()
